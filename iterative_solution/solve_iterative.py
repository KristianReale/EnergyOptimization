from pathlib import Path

import pandas as pd
import numpy as np


import time
import os
import csv

from openpyxl import load_workbook

inputFolder = "./csv/k/"
outputFolderFile = "../../Results/nuovi_dati/greedy/k/greedy_analysis.xlsx"

file_list = [f for f in os.listdir(inputFolder) if
                  f.endswith(
                      ".csv")]
for fIn in file_list:

    #df = pd.read_excel(inputFolder + fIn)

    start = time.time()  # tempo iniziale

    maxCharge = 100
    minCharge = 20
    initChargePercentage = 60
    maxDischargeTimes = 25
    countDischargeTimes = 0

    # === 2️⃣ Assicurati che la colonna 'date' sia di tipo datetime ===
    '''df['date'] = pd.to_datetime(df['date'])
    df["MaxCharge"] = None
    df[maxCharge] = None
    df["MinCharge"] = None
    df[minCharge] = None
    df["InitChargePercentage"] = None
    df[initChargePercentage] = None'''

    # === 3️⃣ Itera sulle righe e calcola i valori mancanti ===
    soc_value = maxCharge * initChargePercentage / 100
    #print(soc_value)
    with open(inputFolder + fIn, 'r') as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader)  # Legge l'intestazione
        counterRow = 0
        dates = []
        productions = []
        consumptions = []
        discharges = []
        charges = []
        feed_ins = []
        from_grids = []
        soc_percents = []
        soc_values = []
        for row in reader:
            date, production, consumption, k = row
            maxDischargeTimes = int(k)
            production = float(production)
            consumption = float(consumption)
            # Esempio: ottieni i valori noti della riga
            '''
            date = row['date']
            discharge = row['Discharge(KW)']
            charge = row['Charge(KW)']
            production = row['Production(KW)']
            consumption = row['Consumption(KW)']
            feed_in = row['Feed-in(KW)']
            from_grid = row['From grid(KW)']
            soc_percent = row['State of Charge (%)']
            '''

            # === 4️⃣ Se un valore è mancante (NaN), calcolalo con la tua formula ===

            discharge = 0
            charge = 0
            feed_in = 0
            from_grid = 0

            if production < consumption:
                discharge = consumption - production
                if soc_value - discharge < minCharge:
                    discharge = soc_value - minCharge
                if discharge > 0:
                    if countDischargeTimes <    maxDischargeTimes:
                        countDischargeTimes += 1
                    else:
                        discharge = 0
                from_grid = consumption - production - discharge
            else:
                charge = production - consumption
                if charge + soc_value > maxCharge:
                    charge = maxCharge - soc_value
                feed_in = production - consumption - charge

            soc_value = soc_value - discharge + charge

            soc_percent = soc_value / maxCharge * 100
            dates.append(date)
            productions.append(production)
            consumptions.append(consumption)
            discharges.append(discharge)
            charges.append(charge)
            feed_ins.append(feed_in)
            from_grids.append(from_grid)
            soc_percents.append(soc_percent)
            soc_values.append(soc_value)

        res_df = pd.DataFrame({
            "date": dates,
            "Discharge(KW)" : discharges,
            "Charge(KW)" : charges,
            "Production(KW)" : productions,
            "Consumption(KW)" : consumptions,
            "Feed - in (KW)" : feed_ins,
            "From grid(KW)" : from_grids,
            "State of Charge(%)": soc_percents,
            "State of Charge(Value)": soc_values
        })


        # === 6️⃣ (Facoltativo) Salva il risultato ===
        end = time.time()  # tempo finale

        mode = "a" if Path(outputFolderFile).exists() else "w"
        sheet_name = fIn.split(".")[0]
        with pd.ExcelWriter(outputFolderFile, mode=mode) as writer:
            res_df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = load_workbook(outputFolderFile)
        ws = wb[sheet_name]
        ws["B26"] = "=SUM(B2:B25)"
        ws["C26"] = "=SUM(C2:C25)"
        ws["D26"] = "=SUM(D2:D25)"
        ws["E26"] = "=SUM(E2:E25)"
        ws["F26"] = "=SUM(F2:F25)"
        ws["G26"] = "=SUM(G2:G25)"
        ws["A28"] = "Max K"
        ws["B28"] = maxDischargeTimes
        ws['A29'] = "Conteggio"
        ws['B29'] = "=COUNTIF(B2:B25,\">0\")"
        wb.save(outputFolderFile)

        print("Tempo di esecuzione:", end - start, "secondi")
        print("Calcolo dei valori mancanti completato.")
